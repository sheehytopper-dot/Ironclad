"""Pure tests for the Step 4 Tenants tab (ui/tabs/tenants_tab.py,
ui/convert.py additions — Phase 5; no browser).

Acceptance (NEXT_STEPS_TO_PHASE5.md Step 4, advisor directive):
1. Round-trip + §25 one-field discrimination on the real goldens; readable
   per-cell errors (field path + offending value, no pydantic/Traceback).
2. Importing the Phase 4 template export of a golden lands the Contractual
   subset identically (flat fields), the Speculative-rows note is present,
   and a malformed row surfaces the Step-7 readable error verbatim.
3. The rollover-generations rows (the D6-amendment Freeport E surface)
   show Freeport's known per-generation values — LC pct 6.75 and renewal
   weight 0.75 — as §25-discriminating literals; a contract generation is
   labeled Contractual with no LC (wrong-field/wrong-generation reads
   fail).
"""
from pathlib import Path

import openpyxl
import pytest

from engine.calc.run import run_property
from engine.export import export_rent_roll
from engine.export.rent_roll_export import RENT_ROLL_COLUMNS
from engine.intake import RentRollImportError, import_rent_roll
from engine.reports import CONTRACTUAL, SPECULATIVE
from ui import convert, state
from ui.tabs import tenants_tab

GOLDEN = Path(__file__).resolve().parents[1] / "golden"
FREEPORT = GOLDEN / "freeport" / "freeport.icprop.json"
CLOROX = GOLDEN / "clorox_northlake" / "clorox_northlake.icprop.json"


@pytest.fixture(scope="module")
def freeport():
    model, error = state.load_model(FREEPORT)
    assert error is None
    return model


@pytest.fixture(scope="module")
def freeport_result(freeport):
    return run_property(freeport)


def _assert_edit_roundtrips_and_discriminates(original, edited, tmp_path):
    assert edited is not None
    assert not state.models_equal(original, edited)
    saved = state.save_model(edited, tmp_path / "edited.icprop.json")
    reloaded, error = state.load_model(saved)
    assert error is None
    assert state.models_equal(edited, reloaded)


class TestLeaseGrid:
    def test_scalar_edit_preserves_nested(self, freeport, tmp_path):
        data = freeport.model_dump(mode="json")
        rows = convert.lease_grid_rows(data["rent_roll"])
        rows[0]["suite"] = "999"
        edited, error = tenants_tab.apply_lease_grid(freeport, rows)
        assert error is None
        # a steps-bearing lease keeps its steps through an unrelated edit
        assert len(edited.rent_roll[1].rent_steps) == len(
            freeport.rent_roll[1].rent_steps)
        _assert_edit_roundtrips_and_discriminates(freeport, edited, tmp_path)

    def test_base_rent_flatten_unflatten(self, freeport, tmp_path):
        data = freeport.model_dump(mode="json")
        rows = convert.lease_grid_rows(data["rent_roll"])
        rows[0]["base_rent_amount"] = 44.44
        edited, error = tenants_tab.apply_lease_grid(freeport, rows)
        assert error is None
        assert edited.rent_roll[0].base_rent.amount == 44.44
        assert (edited.rent_roll[0].base_rent.unit
                == freeport.rent_roll[0].base_rent.unit)  # unit untouched
        _assert_edit_roundtrips_and_discriminates(freeport, edited, tmp_path)

    def test_add_row_uses_template(self, freeport):
        data = freeport.model_dump(mode="json")
        rows = convert.lease_grid_rows(data["rent_roll"])
        rows.append({"tenant_name": "Newco", "area": 2_500.0,
                     "lease_type": "office", "start_date": "2027-01-01",
                     "term_months": 36, "base_rent_amount": 20.0,
                     "base_rent_unit": "dollars_per_area_per_year"})
        edited, error = tenants_tab.apply_lease_grid(freeport, rows)
        assert error is None
        added = edited.rent_roll[-1]
        assert added.tenant_name == "Newco"
        assert added.base_rent.amount == 20.0
        # the template default: valid WITHOUT an MLP link (vacate)
        assert added.upon_expiration.value == "vacate"

    def test_market_without_mlp_readable_error(self, freeport):
        """A new row with upon_expiration=market but no MLP trips the real
        §3.12 cross-field rule, readably."""
        data = freeport.model_dump(mode="json")
        rows = convert.lease_grid_rows(data["rent_roll"])
        rows.append({"tenant_name": "Newco", "area": 2_500.0,
                     "lease_type": "office", "start_date": "2027-01-01",
                     "term_months": 36, "base_rent_amount": 20.0,
                     "base_rent_unit": "dollars_per_area_per_year",
                     "upon_expiration": "market"})
        edited, error = tenants_tab.apply_lease_grid(freeport, rows)
        assert edited is None
        assert "market_leasing_profile" in error
        assert "Traceback" not in error and "pydantic" not in error.lower()

    def test_both_end_and_term_readable_error(self, freeport):
        # Freeport lease 0 (OKI) uses end_date — setting term_months too
        # trips the exactly-one-of validator
        data = freeport.model_dump(mode="json")
        assert data["rent_roll"][0]["end_date"] is not None  # fixture truth
        rows = convert.lease_grid_rows(data["rent_roll"])
        rows[0]["term_months"] = 60
        edited, error = tenants_tab.apply_lease_grid(freeport, rows)
        assert edited is None
        assert "end_date" in error and "term_months" in error
        assert "Traceback" not in error and "pydantic" not in error.lower()

    def test_all_or_nothing(self, freeport):
        before = freeport.model_dump()
        data = freeport.model_dump(mode="json")
        rows = convert.lease_grid_rows(data["rent_roll"])
        rows[0]["area"] = -1
        tenants_tab.apply_lease_grid(freeport, rows)
        assert freeport.model_dump() == before


class TestLeaseDetail:
    def test_full_detail_payload_roundtrips(self, freeport, tmp_path):
        data = freeport.model_dump(mode="json")
        lease = data["rent_roll"][2]
        payload = {
            "rent_steps": [{"month_offset": 12, "date": None, "amount": None,
                            "unit": None, "pct_increase": 3.0}],
            "cpi": {"method": "full_cpi", "index": None, "pct": None,
                    "first_increase_month": "anniversary",
                    "frequency_months": 12, "cap_pct": None,
                    "floor_pct": None},
            "free_rent": {"months": 2.0, "timing": "front",
                          "custom_months": None, "profile": None},
            "security_deposit": {"amount": 2.0, "unit": "months_of_rent",
                                 "refunded_at_expiration": True},
            "recoveries": {**lease["recoveries"], "method": "base_stop",
                           "stop_amount_per_area": 4.5},
        }
        edited, error = tenants_tab.apply_lease_detail(freeport, 2, payload)
        assert error is None
        got = edited.rent_roll[2]
        assert got.cpi.method.value == "full_cpi"
        assert got.free_rent.months == 2.0
        assert got.security_deposit.amount == 2.0
        assert got.recoveries.method.value == "base_stop"
        _assert_edit_roundtrips_and_discriminates(freeport, edited, tmp_path)

    def test_pct_rent_editor_payload(self, freeport, tmp_path):
        payload = {"percentage_rent": {
            "sales_volume": {"amount": 400.0,
                             "unit": "dollars_per_area_per_year",
                             "growth": None},
            "breakpoint": "natural",
            "breakpoint_layers": [{"breakpoint_amount": None, "pct": 6.0}]}}
        edited, error = tenants_tab.apply_lease_detail(freeport, 0, payload)
        assert error is None
        assert edited.rent_roll[0].percentage_rent.breakpoint_layers[0].pct == 6.0
        _assert_edit_roundtrips_and_discriminates(freeport, edited, tmp_path)

    def test_invalid_cpi_readable(self, freeport):
        edited, error = tenants_tab.apply_lease_detail(
            freeport, 0, {"cpi": {"method": "cpi_squared"}})
        assert edited is None
        assert "cpi" in error and "Traceback" not in error

    def test_misc_item_rows_merge(self):
        items = [{"name": "Storage", "amount": 100.0,
                  "unit": "dollars_per_month", "free_rent_abates": True,
                  "timing": {"method": "continuous", "start": None,
                             "end": None, "repeat_months": None,
                             "repeat_every_months": None}}]
        rows = convert.misc_items_to_rows(items)
        rows[0]["amount"] = 150.0
        merged = convert.apply_misc_item_rows(items, rows)
        assert merged[0]["amount"] == 150.0
        assert merged[0]["timing"]["method"] == "continuous"  # preserved


class TestAbsorptionAndStructures:
    def test_absorption_edit_roundtrips(self, freeport, tmp_path):
        data = freeport.model_dump(mode="json")
        rows = convert.absorption_to_rows(data["absorption"])
        rows[0]["total_area"] = 2_400.0
        edited, error = tenants_tab.apply_absorption(freeport, rows)
        assert error is None
        assert edited.absorption[0].total_area == 2_400.0
        _assert_edit_roundtrips_and_discriminates(freeport, edited, tmp_path)

    def test_structure_pool_edit_roundtrips(self, freeport, tmp_path):
        data = freeport.model_dump(mode="json")
        structure = data["recovery_structures"][0]
        pools = [dict(p) for p in structure["pools"]]
        pools[0]["admin_fee_pct"] = 3.0
        edited, error = tenants_tab.apply_recovery_structure(
            freeport, 0, {"name": structure["name"], "pools": pools})
        assert error is None
        assert edited.recovery_structures[0].pools[0].admin_fee_pct == 3.0
        _assert_edit_roundtrips_and_discriminates(freeport, edited, tmp_path)

    def test_add_and_delete_structure(self, freeport):
        added, error = tenants_tab.add_recovery_structure(freeport,
                                                          "Test Structure")
        assert error is None
        assert added.recovery_structures[-1].name == "Test Structure"
        deleted, error = tenants_tab.delete_recovery_structure(
            added, len(added.recovery_structures) - 1)
        assert error is None
        assert len(deleted.recovery_structures) == len(
            freeport.recovery_structures)

    def test_bad_pool_method_readable(self, freeport):
        data = freeport.model_dump(mode="json")
        structure = data["recovery_structures"][0]
        pools = [dict(p) for p in structure["pools"]]
        pools[0]["method"] = "everything"
        edited, error = tenants_tab.apply_recovery_structure(
            freeport, 0, {"name": structure["name"], "pools": pools})
        assert edited is None
        assert "pools.0.method" in error
        assert "Traceback" not in error and "pydantic" not in error.lower()


class TestTemplateImport:
    @pytest.fixture()
    def template(self, freeport_result, tmp_path):
        path = tmp_path / "template.xlsx"
        export_rent_roll(freeport_result, path=path)
        return path

    def test_contractual_subset_lands_identically(self, freeport, template):
        imported = import_rent_roll(template)
        assert len(imported.leases) == len(freeport.rent_roll)  # 29
        edited, error = tenants_tab.apply_imported_rent_roll(
            freeport, imported.leases)
        assert error is None
        for original, got in zip(freeport.rent_roll, edited.rent_roll):
            # the flat §5.2 fields the template carries
            assert got.tenant_name == original.tenant_name
            assert got.area == original.area
            assert got.base_rent.amount == original.base_rent.amount
            assert got.base_rent.unit == original.base_rent.unit
            assert got.status == original.status
            assert len(got.rent_steps) == len(original.rent_steps)

    def test_speculative_note_present(self, template):
        imported = import_rent_roll(template)
        assert len(imported.notes) == 1
        note = imported.notes[0]
        assert "speculative" in note and "ignored" in note
        assert "not intake" in note

    def test_malformed_row_readable_verbatim(self, template):
        wb = openpyxl.load_workbook(template)
        column = RENT_ROLL_COLUMNS.index("lease_type") + 1
        wb["Rent Roll"].cell(row=2, column=column).value = "offce"
        wb.save(template)
        with pytest.raises(RentRollImportError) as excinfo:
            import_rent_roll(template)
        message = str(excinfo.value)
        assert "Rent Roll sheet" in message and "row 2" in message
        assert "'offce'" in message
        assert "office, industrial, retail" in message
        assert "Traceback" not in message
        assert "pydantic" not in message.lower()

    def test_import_with_unknown_mlp_fails_readably(self, freeport_result,
                                                    tmp_path, freeport):
        """Applying imported leases revalidates the WHOLE document — a lease
        referencing an MLP the model doesn't define is refused readably."""
        path = tmp_path / "template.xlsx"
        export_rent_roll(freeport_result, path=path)
        wb = openpyxl.load_workbook(path)
        column = RENT_ROLL_COLUMNS.index("market_leasing_profile") + 1
        wb["Rent Roll"].cell(row=2, column=column).value = "Ghost MLA"
        wb.save(path)
        imported = import_rent_roll(path)          # template itself is valid
        edited, error = tenants_tab.apply_imported_rent_roll(
            freeport, imported.leases)
        assert edited is None
        assert "Ghost MLA" in error or "market_leasing_profile" in error
        assert "Traceback" not in error


class TestRolloverGenerationsPanel:
    """The D6-amendment Freeport E surface (§25-discriminating literals)."""

    def test_freeport_generation_literals(self, freeport_result):
        rows = convert.segments_to_generation_rows(
            freeport_result.segments["Aqore LLC"], CONTRACTUAL, SPECULATIVE)
        assert len(rows) == 3                      # contract + 2 rollovers
        contract, spec1, spec2 = rows
        # the contract generation: Contractual, full weight, NO LC/TI
        assert contract["provenance"] == CONTRACTUAL
        assert contract["renewal_weight"] == 1.0
        assert contract["lc"] == "" and contract["ti"] == ""
        # the rollover generations: the KNOWN Freeport economics —
        # LC 6.75% of rent, renewal weight 0.75, TI 12.5 $/SF (wrong-field
        # or wrong-generation reads fail these)
        for spec in (spec1, spec2):
            assert spec["provenance"] == SPECULATIVE
            assert spec["renewal_weight"] == 0.75
            assert spec["lc"] == "6.75% of rent"
            assert spec["ti"] == "12.5 dollars_per_area"
            assert spec["free_rent_months"] == 5.0
            assert spec["initial_rent_monthly"] is not None

    def test_vacate_chain_has_single_contract_generation(self,
                                                         freeport_result):
        """OKI Data #1 vacates — its chain is one Contractual generation
        (the panel must NOT invent rollovers)."""
        rows = convert.segments_to_generation_rows(
            freeport_result.segments["OKI Data Americas Inc."],
            CONTRACTUAL, SPECULATIVE)
        assert len(rows) == 1
        assert rows[0]["provenance"] == CONTRACTUAL

    def test_absorption_first_term_is_speculative(self, freeport_result):
        """The absorption chain's own first term is labeled Speculative
        (lease.status — the [AE p. 398] rule the reports/export use)."""
        tenant = next(t for t in freeport_result.segments
                      if "lease-up" in t)
        rows = convert.segments_to_generation_rows(
            freeport_result.segments[tenant], CONTRACTUAL, SPECULATIVE)
        assert rows[0]["provenance"] == SPECULATIVE
