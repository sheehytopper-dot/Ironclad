"""Unit tests for the Phase 4 Step 6 Excel export package
(engine/export/package_builder.py, engine/export/rent_roll_export.py; spec §8,
§5.2).

Acceptance (NEXT_STEPS_TO_PHASE4 Step 6): each tab's cell values equal the
report builder's DataFrame EXACTLY — the workbook is read back with openpyxl
and diffed cell-by-cell against the builders. Per the DEVIATIONS §25
standing rule the diff must be capable of failing: ``test_diff_catches_a_
corrupted_cell`` corrupts a written cell and confirms the comparison raises.
The exporter recomputes nothing; any building-area figure it surfaces is the
run's rentable area, never a summed-contract-area (verified on Freeport).
"""
import datetime as dt
import math

import openpyxl
import pytest

from engine.calc.run import run_property
from engine.export import (
    DATA_START_ROW,
    DEFAULT_REPORTS,
    RENT_ROLL_COLUMNS,
    build_package,
    export_report,
    export_rent_roll,
    report_cell_grid,
)
from engine.export.package_builder import _cellify
from engine.reports import cash_flow, Period
from engine.models import (
    AreaMeasures,
    ExpenseItem,
    ExpenseUnit,
    Inflation,
    Lease,
    MoneyRate,
    MoneyUnit,
    PropertyInfo,
    PropertyModel,
    Purchase,
    RentableAreaMode,
    RentStep,
    TimingBasis,
    UponExpiration,
    YearRate,
)
from engine.models.investment import ClosingCost, Loan, LoanAmount, LoanCosts
from engine.models.valuation import (
    DirectCap,
    Resale,
    SensitivityIntervals,
    ValuationInputs,
)

BEGIN = dt.date(2026, 1, 1)
PSF_YR = MoneyUnit.dollars_per_area_per_year
FLAT = Inflation(general_rate=[YearRate(year=1, rate=3.0)],
                 timing_basis=TimingBasis.analysis_year)


def full_model(*, loans=True, valuation=True):
    lease = Lease(tenant_name="Acme Co", suite="100", area=12_000,
                  lease_type="industrial", start_date=BEGIN, term_months=120,
                  base_rent=MoneyRate(amount=10.0, unit=PSF_YR),
                  upon_expiration=UponExpiration.vacate,
                  rent_steps=[RentStep(month_offset=24, amount=11.0,
                                       unit=PSF_YR)])
    kwargs = dict(
        property=PropertyInfo(name="Vista Tower", property_type="industrial",
                              analysis_begin=BEGIN, analysis_term_years=5),
        area_measures=AreaMeasures(
            property_size=12_000, rentable_area_mode=RentableAreaMode.fixed,
            rentable_area_fixed=12_000),
        inflation=FLAT, rent_roll=[lease],
        expenses=[ExpenseItem(name="CAM", amount=20_000.0,
                              unit=ExpenseUnit.dollars_per_year,
                              recoverable=True)])
    if loans:
        kwargs["loans"] = [Loan(name="Mortgage",
                                amount=LoanAmount(value=600_000.0),
                                term_months=360, rate=6.0,
                                amortization="fully_amortizing",
                                loan_costs=LoanCosts(points_pct=1.0))]
    if valuation:
        kwargs["purchase"] = Purchase(
            price=1_250_000.0,
            closing_costs=[ClosingCost(name="Legal", amount=25_000.0)])
        kwargs["valuation"] = ValuationInputs(
            discount_rate=8.0, direct_cap=DirectCap(cap_rate=8.0),
            resale=Resale(method="cap_noi_current_year", exit_cap_rate=8.0,
                          selling_costs_pct=3.0),
            sensitivity_intervals=SensitivityIntervals(
                discount_rate_step=1.0, cap_rate_step=1.0, count=5))
    return PropertyModel(**kwargs)


def _cell_eq(a, b) -> bool:
    if a is None and b is None:
        return True
    if isinstance(a, (int, float)) and isinstance(b, (int, float)) \
            and not isinstance(a, bool) and not isinstance(b, bool):
        return math.isclose(float(a), float(b), rel_tol=1e-9, abs_tol=1e-6)
    return str(a) == str(b)


def assert_tab_matches_grid(worksheet, grid) -> None:
    """Cell-by-cell: the written data region equals ``grid`` exactly. Raises
    AssertionError (with the first mismatch) otherwise — so it CAN fail."""
    for i, row_cells in enumerate(grid):
        for j, expected in enumerate(row_cells):
            got = worksheet.cell(row=DATA_START_ROW + 1 + i,
                                 column=j + 1).value
            assert _cell_eq(_cellify(got), expected), (
                f"cell [{i},{j}] on {worksheet.title!r}: "
                f"wrote {got!r}, expected {expected!r}")


@pytest.fixture(scope="module")
def model():
    return full_model()


@pytest.fixture(scope="module")
def result(model):
    return run_property(model)


@pytest.fixture(scope="module")
def workbook_path(result, model, tmp_path_factory):
    path = tmp_path_factory.mktemp("export") / "package.xlsx"
    build_package(result, model, path=path, scenario="Base",
                  timestamp="2026-07-14 12:00")
    return path


class TestPackageCellByCell:
    def test_all_default_tabs_written(self, workbook_path):
        wb = openpyxl.load_workbook(workbook_path, read_only=True)
        assert wb.sheetnames == [
            "Executive Summary", "Annual Cash Flow", "Monthly Cash Flow",
            "Lease Summary", "Lease Expiration", "IRR Matrix", "Value Matrix",
            "Present Value", "Recovery Audit", "Loan Amortization",
            "Assumptions"]

    def test_every_tab_equals_its_builder_dataframe(self, workbook_path,
                                                    result, model):
        wb = openpyxl.load_workbook(workbook_path, data_only=True)
        for spec in DEFAULT_REPORTS:
            if not spec.applies(result, model):
                continue
            grid = report_cell_grid(spec.build(result, model))
            assert_tab_matches_grid(wb[spec.tab[:31]], grid)


class TestDiscrimination:
    """DEVIATIONS §25: prove the cell-by-cell check can fail. Corrupt a
    written cell, re-read, and confirm the comparison raises."""

    def test_diff_catches_a_corrupted_cell(self, result, model, tmp_path):
        path = tmp_path / "corrupt.xlsx"
        build_package(result, model, path=path, timestamp="t")
        grid = report_cell_grid(cash_flow(
            result, period=Period.fiscal,
            fiscal_year_end_month=model.property.fiscal_year_end_month,
            analysis_begin=model.property.analysis_begin))

        # sanity: the clean workbook matches
        wb = openpyxl.load_workbook(path, data_only=True)
        assert_tab_matches_grid(wb["Annual Cash Flow"], grid)

        # corrupt one written data cell (a numeric NOI-area cell), save
        wb_edit = openpyxl.load_workbook(path)
        ws = wb_edit["Annual Cash Flow"]
        target = ws.cell(row=DATA_START_ROW + 3, column=2)
        target.value = (target.value or 0) + 12345.0
        wb_edit.save(path)

        # the comparison must now FAIL on that cell
        wb2 = openpyxl.load_workbook(path, data_only=True)
        with pytest.raises(AssertionError):
            assert_tab_matches_grid(wb2["Annual Cash Flow"], grid)


class TestApplicabilityAndBuildingArea:
    def test_skips_loan_and_valuation_tabs_when_absent(self, tmp_path):
        bare = full_model(loans=False, valuation=False)
        res = run_property(bare)
        path = tmp_path / "bare.xlsx"
        build_package(res, bare, path=path, timestamp="t")
        names = set(openpyxl.load_workbook(path, read_only=True).sheetnames)
        assert "Loan Amortization" not in names
        assert "IRR Matrix" not in names and "Value Matrix" not in names
        assert "Present Value" not in names
        # the always-on tabs are still there
        assert {"Executive Summary", "Annual Cash Flow", "Lease Summary",
                "Assumptions"} <= names

    def test_exported_building_area_is_rentable_not_summed_contract(self,
                                                                    tmp_path):
        """DEVIATIONS §25 at the export layer, on a discriminating fixture:
        Freeport's Executive Summary tab surfaces rentable 123,099, NOT the
        summed contract area 128,087 (a 4,988 SF gap)."""
        from pathlib import Path
        from engine.models.io import load_property
        fixture = (Path(__file__).resolve().parents[1] / "golden" /
                   "freeport" / "freeport.icprop.json")
        fp_model = load_property(fixture)
        res = run_property(fp_model)
        summed_contract = sum(
            next(s for s in segs if not s.speculative).area
            for segs in res.segments.values())
        assert summed_contract == pytest.approx(128_087.0)  # the wrong answer

        path = tmp_path / "freeport.xlsx"
        build_package(res, fp_model, path=path, timestamp="t")
        ws = openpyxl.load_workbook(path, data_only=True)["Executive Summary"]
        reported = None
        for row in ws.iter_rows(values_only=True):
            if row and row[0] == "Rentable Area (SF)":
                reported = row[1]
        assert reported == pytest.approx(123_099.0)          # rentable
        assert reported != pytest.approx(summed_contract)    # NOT summed


class TestSingleReportExport:
    def test_export_report_round_trips_grid(self, result, model, tmp_path):
        report = cash_flow(result, period=Period.annual,
                           analysis_begin=BEGIN)
        path = tmp_path / "one.xlsx"
        tab = export_report(report, path=path, tab="Cash Flow")
        wb = openpyxl.load_workbook(path, data_only=True)
        assert wb.sheetnames == [tab]
        assert_tab_matches_grid(wb[tab], report_cell_grid(report))


class TestRentRollExport:
    def test_template_layout_and_values(self, result, tmp_path):
        path = tmp_path / "rr.xlsx"
        counts = export_rent_roll(result, path=path)
        assert counts["Rent Roll"] == 1 and counts["Rent Steps"] == 1
        assert counts["Contractual"] == 1 and counts["Speculative"] == 0
        wb = openpyxl.load_workbook(path)
        assert wb.sheetnames == ["Rent Roll", "Rent Steps", "Misc Items"]
        header = [c.value for c in wb["Rent Roll"][1]]
        assert header == RENT_ROLL_COLUMNS
        row = [c.value for c in wb["Rent Roll"][2]]
        by = dict(zip(header, row))
        assert by["status"] == "Contractual"          # provenance
        assert by["lease_status"] == "contract"       # §3.12 status
        assert by["tenant_name"] == "Acme Co"
        assert by["area"] == 12_000
        assert by["base_rent_amount"] == 10.0
        assert by["base_rent_unit"] == "dollars_per_area_per_year"
        # the rent step round-trips to the companion sheet
        step = [c.value for c in wb["Rent Steps"][2]]
        assert step[0] == "Acme Co" and step[1] == 11.0
