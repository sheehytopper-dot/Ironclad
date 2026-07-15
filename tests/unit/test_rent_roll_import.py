"""Unit tests for the rent-roll importer + provenance round-trip
(engine/intake/rent_roll_import.py, engine/export/rent_roll_export.py;
spec §5.2 / §5.4).

Acceptance (report/export polish pass, 2026-07-15):
1. ROUND-TRIP: the **Contractual** subset round-trips exactly (every flat
   field); **Speculative** rows appear in the export but produce NO lease on
   import.
2. §25 DISCRIMINATION: relabelling a Contractual row to Speculative DROPS it
   from the imported leases (the filter is real and can fail); a Speculative
   row never becomes a Lease.
3. READABLE ERRORS (Step 7, preserved): a malformed Contractual row yields a
   plain, row-level message (sheet/row/column/value/fix), never a pydantic
   dump; blank required cells are readable too.
4. The ignored Speculative rows are reported in a readable ``note`` — not a
   silent skip.
"""
import datetime as dt

import openpyxl
import pytest

from engine.calc.run import run_property
from engine.export import export_rent_roll
from engine.export.rent_roll_export import RENT_ROLL_COLUMNS
from engine.intake import (
    ImportResult,
    RentRollImportError,
    import_rent_roll,
    import_rent_roll_csv,
)
from engine.models import (
    AreaMeasures,
    Inflation,
    Lease,
    LeaseStatus,
    MarketLeasingProfile,
    MoneyRate,
    MoneyUnit,
    PctOfNew,
    PropertyInfo,
    PropertyModel,
    RentableAreaMode,
    RentStep,
    TimingBasis,
    UponExpiration,
    YearRate,
)
from engine.models.profiles import MiscItemSpec

PSF = MoneyUnit.dollars_per_area_per_year
BEGIN = dt.date(2026, 1, 1)


def _model():
    """Three Contractual leases — Acme (steps + misc), Roller (rolls over →
    Speculative generations), Beta (MTM) — on a fixed 20k building."""
    profile = MarketLeasingProfile(
        name="M", term_months=24, renewal_probability=50.0, months_vacant=2.0,
        market_base_rent_new=MoneyRate(amount=15.0, unit=PSF),
        market_base_rent_renew=PctOfNew(pct_of_new=100.0),
        upon_expiration=UponExpiration.market, term_growth=False)
    acme = Lease(tenant_name="Acme Co", suite="100", external_id="EXT1",
                 area=12_000, lease_type="office", start_date=BEGIN,
                 term_months=60, base_rent=MoneyRate(amount=25.0, unit=PSF),
                 upon_expiration=UponExpiration.vacate, notes="anchor",
                 rent_steps=[RentStep(month_offset=24, amount=27.0, unit=PSF)],
                 miscellaneous_items=[MiscItemSpec(
                     name="Storage", amount=100.0,
                     unit=MoneyUnit.dollars_per_month, free_rent_abates=True)])
    roller = Lease(tenant_name="Roller LLC", suite="200", area=5_000,
                   lease_type="office", start_date=BEGIN, term_months=24,
                   base_rent=MoneyRate(amount=20.0, unit=PSF),
                   market_leasing_profile="M",
                   upon_expiration=UponExpiration.market)
    beta = Lease(tenant_name="Beta MTM", suite="300", area=3_000,
                 lease_type="retail", start_date=BEGIN, term_months=60,
                 base_rent=MoneyRate(amount=30.0, unit=PSF), status="mtm",
                 upon_expiration=UponExpiration.vacate)
    return PropertyModel(
        property=PropertyInfo(name="T", property_type="office",
                              analysis_begin=BEGIN, analysis_term_years=5),
        area_measures=AreaMeasures(property_size=20_000,
                                   rentable_area_mode=RentableAreaMode.fixed,
                                   rentable_area_fixed=20_000),
        inflation=Inflation(general_rate=[YearRate(year=1, rate=0.0)],
                            timing_basis=TimingBasis.analysis_year),
        market_leasing_profiles=[profile], rent_roll=[acme, roller, beta])


def _col(name):
    return RENT_ROLL_COLUMNS.index(name) + 1


def _set_cell(path, row, col_name, value):
    wb = openpyxl.load_workbook(path)
    wb["Rent Roll"].cell(row=row, column=_col(col_name)).value = value
    wb.save(path)


def _assert_contractual_roundtrip(model, imported) -> None:
    """The Contractual subset (model.rent_roll) round-trips exactly."""
    by_name = {l.tenant_name: l for l in imported}
    assert set(by_name) == {l.tenant_name for l in model.rent_roll}
    for o in model.rent_roll:
        g = by_name[o.tenant_name]
        assert g.suite == o.suite and g.external_id == o.external_id
        assert g.area == o.area and g.lease_type == o.lease_type
        assert g.status == o.status               # incl. Beta's mtm
        assert g.start_date == o.start_date and g.end_date == o.end_date
        assert g.term_months == o.term_months
        assert g.base_rent.amount == o.base_rent.amount
        assert g.base_rent.unit == o.base_rent.unit
        assert g.upon_expiration == o.upon_expiration
        assert g.notes == o.notes
        assert len(g.rent_steps) == len(o.rent_steps)
        for so, sg in zip(o.rent_steps, g.rent_steps):
            assert (sg.amount, sg.pct_increase, sg.unit, sg.month_offset) == (
                so.amount, so.pct_increase, so.unit, so.month_offset)
        assert len(g.miscellaneous_items) == len(o.miscellaneous_items)
        for mo, mg in zip(o.miscellaneous_items, g.miscellaneous_items):
            assert (mg.name, mg.amount, mg.unit, mg.free_rent_abates) == (
                mo.name, mo.amount, mo.unit, mo.free_rent_abates)


@pytest.fixture
def exported(tmp_path):
    model = _model()
    result = run_property(model)
    path = tmp_path / "rent_roll.xlsx"
    export_rent_roll(result, path=path)
    return model, result, path


class TestContractualRoundTrip:
    def test_contractual_subset_round_trips(self, exported):
        model, _result, path = exported
        res = import_rent_roll(path)
        assert isinstance(res, ImportResult)
        assert len(res.leases) == 3            # only the Contractual rows
        _assert_contractual_roundtrip(model, res.leases)

    def test_export_carries_speculative_rows(self, exported):
        _model, _result, path = exported
        wb = openpyxl.load_workbook(path)
        provs = [r[0].value for r in wb["Rent Roll"].iter_rows(min_row=2)]
        assert provs.count("Contractual") == 3
        assert provs.count("Speculative") == 2   # Roller's two rollovers

    def test_speculative_rows_ignored_with_readable_note(self, exported):
        _model, _result, path = exported
        res = import_rent_roll(path)
        assert len(res.notes) == 1
        note = res.notes[0]
        assert "2 speculative" in note and "ignored" in note
        assert "not intake" in note and "rows 4, 5" in note  # the row list


class TestProvenanceDiscrimination:
    """§25: the Contractual/Speculative filter is real and can fail."""

    def test_relabelling_contractual_to_speculative_drops_it(self, exported):
        model, _result, path = exported
        # Acme is Contractual on row 2 — relabel its provenance to Speculative
        _set_cell(path, 2, "status", "Speculative")
        leases = import_rent_roll(path).leases
        names = {l.tenant_name for l in leases}
        assert "Acme Co" not in names          # dropped — the filter is real
        assert names == {"Roller LLC", "Beta MTM"}
        assert len(leases) == 2

    def test_speculative_row_never_becomes_a_lease(self, exported):
        _model, _result, path = exported
        leases = import_rent_roll(path).leases
        # Roller appears 3× in the sheet (1 Contractual + 2 Speculative) but
        # exactly ONCE as an imported lease
        assert sum(1 for l in leases if l.tenant_name == "Roller LLC") == 1

    def test_blank_provenance_stays_contractual(self, exported):
        model, _result, path = exported
        _set_cell(path, 2, "status", None)     # blank Acme's provenance
        leases = import_rent_roll(path).leases
        assert "Acme Co" in {l.tenant_name for l in leases}
        assert len(leases) == 3                # still imported

    @pytest.mark.parametrize("col,new_value", [
        ("area", 99_999.0),
        ("lease_type", "industrial"),
        ("base_rent_amount", 12.5),
        ("base_rent_unit", "dollars_per_year"),
        ("term_months", 48),
        ("lease_status", "speculative"),
        ("notes", "changed"),
    ])
    def test_corrupting_a_contractual_field_breaks_roundtrip(self, exported,
                                                             col, new_value):
        model, _result, path = exported
        _set_cell(path, 2, col, new_value)     # corrupt Acme (row 2)
        with pytest.raises(AssertionError):
            _assert_contractual_roundtrip(model, import_rent_roll(path).leases)


class TestReadableErrorsPreserved:
    def _import_err(self, exported, edits):
        _model, _result, path = exported
        for row, col, value in edits:
            _set_cell(path, row, col, value)
        with pytest.raises(RentRollImportError) as excinfo:
            import_rent_roll(path)
        return excinfo.value

    def test_bad_enum_names_row_column_value_fix(self, exported):
        exc = self._import_err(exported, [(2, "lease_type", "offce")])
        assert any("Rent Roll sheet" in m and "row 2" in m and "lease_type" in m
                   and "'offce'" in m and "office, industrial, retail" in m
                   for m in exc.errors)

    def test_blank_required_field_is_readable_not_pydantic(self, exported):
        exc = self._import_err(exported, [(2, "base_rent_amount", None)])
        assert any("base_rent_amount" in m and "required value is missing" in m
                   and "row 2" in m for m in exc.errors)
        text = str(exc)
        assert "errors.pydantic.dev" not in text
        assert "pydantic" not in text.lower()
        assert "Traceback" not in text

    def test_missing_required_column_named(self, exported):
        _model, _result, path = exported
        wb = openpyxl.load_workbook(path)
        wb["Rent Roll"].cell(row=1, column=_col("area")).value = None
        wb.save(path)
        with pytest.raises(RentRollImportError) as excinfo:
            import_rent_roll(path)
        assert any("area" in m and "missing" in m for m in excinfo.value.errors)

    def test_blank_lease_status_defaults_to_contract(self, exported):
        _model, _result, path = exported
        _set_cell(path, 2, "lease_status", None)   # Acme's §3.12 status blank
        leases = import_rent_roll(path).leases
        acme = next(l for l in leases if l.tenant_name == "Acme Co")
        assert acme.status == LeaseStatus.contract  # schema default

    def test_error_surface_is_not_a_stack_trace(self, exported):
        exc = self._import_err(exported, [(2, "lease_type", "bad")])
        text = str(exc)
        assert "could not be imported" in text and "- Rent Roll sheet" in text
        assert "pydantic" not in text.lower() and "Traceback" not in text


class TestCsvRoundTrip:
    def test_csv_contractual_subset_round_trips(self, tmp_path):
        import csv
        model = _model()
        rr = tmp_path / "rr.csv"
        with open(rr, "w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            w.writerow(RENT_ROLL_COLUMNS)
            for lease in model.rent_roll:
                w.writerow([
                    "Contractual", lease.tenant_name, lease.suite or "",
                    lease.external_id or "", lease.status.value,
                    lease.lease_type.value, lease.area,
                    str(lease.start_date),
                    str(lease.end_date) if lease.end_date else "",
                    lease.term_months if lease.term_months else "",
                    lease.base_rent.amount, lease.base_rent.unit.value,
                    lease.upon_expiration.value,
                    lease.market_leasing_profile or "", lease.notes or ""])
            # one Speculative row that must be ignored
            w.writerow(["Speculative", "Roller LLC (proj)", "200", "",
                        "speculative", "office", 5000, "2028-02-01", "", 24,
                        18.0, "dollars_per_year", "market", "M", ""])
        steps = tmp_path / "steps.csv"
        with open(steps, "w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            w.writerow(["tenant_name", "amount", "unit", "pct_increase",
                        "date", "month_offset"])
            w.writerow(["Acme Co", 27.0, PSF.value, "", "", 24])
        misc = tmp_path / "misc.csv"
        with open(misc, "w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            w.writerow(["tenant_name", "name", "amount", "unit",
                        "free_rent_abates"])
            w.writerow(["Acme Co", "Storage", 100.0, "dollars_per_month",
                        "true"])
        res = import_rent_roll_csv(rr, steps_path=steps, misc_path=misc)
        assert len(res.leases) == 3            # the Speculative row ignored
        assert res.notes and "1 speculative" in res.notes[0]
        _assert_contractual_roundtrip(model, res.leases)
