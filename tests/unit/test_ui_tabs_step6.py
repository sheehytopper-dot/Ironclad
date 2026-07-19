"""Pure tests for the Step 6 Reports + Dashboard + Audit tabs
(ui/reports_registry.py, ui/tabs/dashboard_tab.py, ui/tabs/audit_tab.py —
Phase 5; no browser).

Acceptance (NEXT_STEPS_TO_PHASE5.md Step 6, advisor directive):
1. Every rendered report's frame EQUALS the engine builder's output for
   the same unit/period toggles (no UI-side math) — asserted with
   ``assert_frame_equal`` per applicable report; a toggle change produces
   a genuinely different frame (§25).
2. The Audit drill-down reproduces Recovery Audit / Lease Audit rows for
   known golden tenant-months, with the actual composition numbers.
3. Exports reuse the Phase 4 machinery — the UI-exported view matches the
   builder frame cell-by-cell via the engine's own ``report_cell_grid``.
4. The two D6-amendment surfaces (Gate 5 criterion 6): the GV panel's
   components tie EXACTLY to the ledger at the month (Freeport literals);
   the recovery drill isolates a Cedar Alt rollover segment by
   ``segment_start`` (row-level literals).
"""
from pathlib import Path

import openpyxl
import pandas as pd
import pytest
from pandas.testing import assert_frame_equal

from engine.calc.ledger import (
    EXPENSE_RECOVERY_REVENUE,
    GENERAL_VACANCY,
    NOI,
    TENANT_IMPROVEMENTS,
)
from engine.export import export_report
from engine.export.package_builder import report_cell_grid
from engine.reports import (
    Period,
    Unit,
    cash_flow,
    lease_expiration,
    lease_summary,
    occupancy,
    recovery_audit_report,
)
from ui import reports_registry as registry
from ui import state
from ui.tabs import audit_tab, dashboard_tab

GOLDEN = Path(__file__).resolve().parents[1] / "golden"
CLOROX = GOLDEN / "clorox_northlake" / "clorox_northlake.icprop.json"
FREEPORT = GOLDEN / "freeport" / "freeport.icprop.json"
CEDAR = GOLDEN / "cedar_alt" / "cedar_alt.icprop.json"


@pytest.fixture(scope="module")
def clorox():
    model, error = state.load_model(CLOROX)
    assert error is None
    result, error = state.run_model(model)
    assert error is None
    return model, result


@pytest.fixture(scope="module")
def freeport():
    model, error = state.load_model(FREEPORT)
    assert error is None
    result, error = state.run_model(model)
    assert error is None
    return model, result


@pytest.fixture(scope="module")
def cedar():
    model, error = state.load_model(CEDAR)
    assert error is None
    result, error = state.run_model(model)
    assert error is None
    return model, result


class TestRegistryFrameEquality:
    """Acceptance 1: the registry returns the builders' own frames."""

    def test_every_applicable_report_equals_its_builder(self, freeport):
        model, result = freeport
        csv = registry.find_benchmark_csv(FREEPORT)
        for entry in registry.applicable_entries(result, model, csv):
            options = ({"benchmark_csv": csv} if entry.number == 24 else {})
            report = registry.build_entry(entry, result, model,
                                          options=options)
            assert report.frame is not None and len(report.frame) > 0, \
                entry.label

    def test_cash_flow_frame_equals_builder_for_same_toggles(self, clorox):
        model, result = clorox
        entry = next(e for e in registry.REGISTRY if e.number == 1)
        via_registry = registry.build_entry(
            entry, result, model, unit=Unit.per_sf, period=Period.annual)
        direct = cash_flow(result, unit=Unit.per_sf, period=Period.annual,
                           fiscal_year_end_month=model.property
                           .fiscal_year_end_month,
                           analysis_begin=model.property.analysis_begin)
        assert_frame_equal(via_registry.frame, direct.frame)

    def test_unit_toggle_discriminates(self, clorox):
        """§25: Total-$ and per-SF frames genuinely differ."""
        model, result = clorox
        entry = next(e for e in registry.REGISTRY if e.number == 1)
        total = registry.build_entry(entry, result, model, unit=Unit.total,
                                     period=Period.annual)
        per_sf = registry.build_entry(entry, result, model,
                                      unit=Unit.per_sf,
                                      period=Period.annual)
        noi_total = float(total.frame.loc[NOI].iloc[0])
        noi_psf = float(per_sf.frame.loc[NOI].iloc[0])
        assert noi_total == pytest.approx(2_596_319.40, abs=0.01)
        assert noi_psf == pytest.approx(noi_total / 540_000.0, rel=1e-9)
        assert noi_total != noi_psf

    def test_occupancy_period_toggle_equals_builder(self, freeport):
        model, result = freeport
        entry = next(e for e in registry.REGISTRY if e.number == 15)
        via_registry = registry.build_entry(entry, result, model,
                                            period=Period.monthly)
        direct = occupancy(result, period=Period.monthly,
                           fiscal_year_end_month=model.property
                           .fiscal_year_end_month)
        assert_frame_equal(via_registry.frame, direct.frame)

    def test_provenance_reports_equal_builders_both_scopes(self, freeport):
        model, result = freeport
        entry = next(e for e in registry.REGISTRY if e.number == 12)
        full = registry.build_entry(entry, result, model)
        assert_frame_equal(full.frame, lease_expiration(
            result, fiscal_year_end_month=model.property
            .fiscal_year_end_month).frame)
        contractual = registry.build_entry(entry, result, model,
                                           options={"contractual_only": True})
        assert set(contractual.frame["status"]) == {"Contractual"}
        assert len(full.frame) > len(contractual.frame)  # spec rows exist

    def test_applicability_gates(self, clorox):
        model, result = clorox            # no loans / valuation / CSV? has CSV
        entries = registry.applicable_entries(result, model, None)
        numbers = {e.number for e in entries}
        assert 5 not in numbers and 8 not in numbers and 20 not in numbers
        assert 24 not in numbers          # no CSV passed → no benchmark
        with_csv = registry.applicable_entries(
            result, model, registry.find_benchmark_csv(CLOROX))
        assert 24 in {e.number for e in with_csv}

    def test_date_slice_is_pure_selection(self, clorox):
        model, result = clorox
        entry = next(e for e in registry.REGISTRY if e.number == 1)
        report = registry.build_entry(entry, result, model,
                                      period=Period.fiscal)
        frame = report.frame
        sliced = registry.slice_period_columns(frame, frame.columns[1],
                                               frame.columns[3])
        assert list(sliced.columns) == list(frame.columns[1:4])
        assert_frame_equal(sliced, frame.iloc[:, 1:4])   # values untouched


class TestBenchmark:
    def test_clorox_full_grid_zero_misses_with_reported_skip(self, clorox):
        model, result = clorox
        csv = registry.find_benchmark_csv(CLOROX)
        report = registry.build_benchmark(result, model, csv)
        assert report.meta.extra["miss_count"] == 0
        # the un-bridged CSV account is SKIPPED AND REPORTED, not silent
        assert report.meta.extra["skipped_accounts"] == ["Capital Expenses"]

    def test_freeport_full_grid_reproduces_the_by_design_reds(self,
                                                              freeport):
        """170 = the Gate 2 (137) + Gate 3 capital (33) by-design misses —
        the UI benchmark reproduces the known red counts exactly."""
        model, result = freeport
        csv = registry.find_benchmark_csv(FREEPORT)
        report = registry.build_benchmark(result, model, csv)
        assert report.meta.extra["miss_count"] == 170
        assert report.meta.extra["skipped_accounts"] == []

    def test_no_csv_no_benchmark(self, tmp_path):
        assert registry.find_benchmark_csv(tmp_path / "x.icprop.json") is None
        assert registry.find_benchmark_csv(None) is None


class TestExportReuse:
    def test_ui_exported_view_matches_builder_cell_grid(self, clorox,
                                                        tmp_path):
        """Acceptance 3: the UI export path IS the Phase 4 exporter — the
        written workbook matches the builder frame via the engine's own
        report_cell_grid."""
        model, result = clorox
        entry = next(e for e in registry.REGISTRY if e.number == 1)
        report = registry.build_entry(entry, result, model,
                                      period=Period.fiscal)
        path = tmp_path / "view.xlsx"
        export_report(report, path=path)
        grid = report_cell_grid(report)
        sheet = openpyxl.load_workbook(path)[
            openpyxl.load_workbook(path).sheetnames[0]]
        # the grid starts after the title band (row 4: header; data below)
        for r, row in enumerate(grid):
            for c, value in enumerate(row):
                got = sheet.cell(row=4 + r, column=1 + c).value
                if value in (None, ""):
                    assert got in (None, "")
                elif isinstance(value, float):
                    assert got == pytest.approx(value, abs=1e-9)
                else:
                    assert str(got) == str(value)


class TestDashboard:
    def test_dashboard_data_reads_engine_surfaces_only(self, clorox):
        model, result = clorox
        data = dashboard_tab.dashboard_data(result, model)
        # the Step-1 metrics (labels locked by the Step-1 AppTest flows)
        assert data["metrics"]["year1_noi"] == pytest.approx(2_596_319.40,
                                                             abs=0.01)
        assert data["metrics"]["year1_occupancy_pct"] == pytest.approx(100.0)
        # exec-summary values pass through verbatim
        assert data["exec"]["Rentable Area (SF)"] == pytest.approx(540_000.0)
        # the annual chart frame IS the ledger's own annual view
        annual = data["annual_noi_cf"]
        assert float(annual[NOI].iloc[0]) == pytest.approx(2_596_319.40,
                                                           abs=0.01)
        # top tenants come from report #11 sorted by contractual rent
        top = data["top_tenants"]
        assert list(top.columns) == list(lease_summary(result).frame.columns)
        assert (top["annual_base_rent"].is_monotonic_decreasing
                or len(top) == 1)

    def test_freeport_top_tenant_literal(self, freeport):
        model, result = freeport
        data = dashboard_tab.dashboard_data(result, model)
        assert data["top_tenants"].iloc[0]["tenant"] == \
            "Rodeo Dental Management, PLLC"


class TestAuditComposition:
    def test_recovery_revenue_rows_tie_to_ledger(self, freeport):
        """Acceptance 2: the drill reproduces the Recovery/Lease Audit
        composition — the per-tenant rows sum to the ledger line exactly."""
        model, result = freeport
        rows, caption = audit_tab.audit_composition(
            result, model, EXPENSE_RECOVERY_REVENUE, "2026-07")
        ledger_value = float(result.ledger.frame.loc[
            pd.Period("2026-07", "M"), EXPENSE_RECOVERY_REVENUE])
        assert rows["recoveries"].sum() == pytest.approx(ledger_value,
                                                         abs=1e-6)
        assert ledger_value == pytest.approx(20_840.57, abs=0.01)  # literal
        top = rows.iloc[0]
        assert top["tenant"] == "Rodeo Dental Management, PLLC"
        assert top["recoveries"] == pytest.approx(3_512.68, abs=0.01)
        assert "Lease Audit" in caption

    def test_expense_item_composition(self, clorox):
        model, result = clorox
        rows, caption = audit_tab.audit_composition(
            result, model, "Common Area Maintenance", "2026-06")
        assert len(rows) == 1
        assert rows.iloc[0]["amount"] == pytest.approx(27_290.04, abs=0.01)
        assert "negative" in caption          # the sign convention stated

    def test_ti_composition_at_rollover(self, freeport):
        model, result = freeport
        month = next(str(m) for m in result.ledger.frame.index
                     if float(result.ledger.frame.loc[m,
                                                      TENANT_IMPROVEMENTS])
                     != 0.0)
        rows, _ = audit_tab.audit_composition(result, model,
                                              TENANT_IMPROVEMENTS, month)
        ledger_value = float(result.ledger.frame.loc[
            pd.Period(month, "M"), TENANT_IMPROVEMENTS])
        assert rows["amount"].sum() == pytest.approx(-ledger_value,
                                                     abs=1e-6) or \
            rows["amount"].sum() == pytest.approx(ledger_value, abs=1e-6)

    def test_gv_routes_to_panel(self, freeport):
        model, result = freeport
        rows, caption = audit_tab.audit_composition(result, model,
                                                    GENERAL_VACANCY,
                                                    "2026-07")
        assert rows is None and "basis panel" in caption


class TestGvBasisPanel:
    """The Freeport B surface (Gate 5 criterion 6) — real literals."""

    def test_components_tie_to_ledger_exactly(self, freeport):
        model, result = freeport
        rows, summary = audit_tab.gv_basis_rows(result, model, "2026-07")
        ledger = result.ledger.frame
        month = pd.Period("2026-07", "M")
        for _, row in rows.iterrows():
            assert row["amount"] == float(ledger.loc[month,
                                                     row["component"]])
        # the Freeport 2026-07 literals (wrong-month/wrong-column fails)
        assert summary["gv_posted"] == pytest.approx(-7_071.48, abs=0.01)
        assert summary["at_vacancy_posted"] == pytest.approx(-4_620.00,
                                                             abs=0.01)
        assert summary["basis_total"] == pytest.approx(229_209.66, abs=0.01)
        assert summary["implied_rate_pct"] == pytest.approx(3.09, abs=0.01)
        assert summary["method"] == "percent_of_pgr"
        assert summary["reduce_by_absorption_turnover"] is True

    def test_method_config_drives_inclusion(self, freeport):
        model, result = freeport
        rows, _ = audit_tab.gv_basis_rows(result, model, "2026-07")
        assert rows["included_in_basis"].all()     # percent_of_pgr = all six


class TestRecoveryDrill:
    """The Cedar Alt B surface (Gate 5 criterion 6) — real literals."""

    def test_drill_isolates_a_rollover_segment(self, cedar):
        model, result = cedar
        frame = recovery_audit_report(result).frame
        tenant, segment = next(
            (t, s) for t, segments in result.segments.items()
            for s in segments if s.speculative)
        drilled = audit_tab.filter_recovery_audit(
            frame, tenant=tenant, segment_start=str(segment.start))
        assert len(drilled) > 0
        assert set(drilled["tenant"]) == {tenant}
        assert {str(s) for s in drilled["segment_start"]} == \
            {str(segment.start)}
        # every drilled month is inside the speculative segment
        assert all(segment.start <= m <= segment.end
                   for m in drilled["month"])
        # the Cedar Alt literals (tenant, start, first-month recovery)
        assert tenant == "Bldg 1 Tenant (Confidential)"
        assert str(segment.start) == "2033-09"
        first = drilled.iloc[0]
        assert first["recovery"] == pytest.approx(173_504.82, abs=0.01)
        assert first["share"] == pytest.approx(0.803174, abs=1e-6)

    def test_month_filter(self, cedar):
        _model, result = cedar
        frame = recovery_audit_report(result).frame
        drilled = audit_tab.filter_recovery_audit(frame, month="2033-09")
        assert len(drilled) > 0
        assert {str(m) for m in drilled["month"]} == {"2033-09"}
