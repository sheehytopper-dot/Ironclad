"""Rent-roll template importer (Phase 4 Step 7; spec §5.2 / §5.4)
[import validation concept: AE pp. 62, 171].

Reads the §5.2 rent-roll template that ``engine/export/rent_roll_export.py``
writes — a **Rent Roll** sheet (one row per lease, flat §3.12 fields) plus
**Rent Steps** and **Misc Items** companion sheets keyed by tenant — and
reconstructs the ``list[Lease]`` (the rent-roll portion of a
``PropertyModel``), validating through the §3 pydantic models. ``.xlsx``
and ``.csv`` are both supported (CSV: the Rent Roll rows, with optional
companion CSVs for steps / misc items).

Errors are **translated for a non-programmer** (spec §5.4): every message
names the sheet, the spreadsheet row, the column, the offending value, and
what a valid value looks like — never a pydantic stack trace. All rows are
validated and *every* error is collected before raising, so the user sees
the whole list at once (:class:`RentRollImportError`).

The flat template carries the scalar lease fields (names, area, dates/term,
base rent amount+unit, status, expiration, MLP reference) plus rent steps
and the misc-item name/amount/unit/abatement. Nested structures (free-rent
profiles, recovery assignments, percentage rent, security deposits, leasing
costs, and misc-item timing/inflation/limits) live in the JSON document
(§5.1) and are out of the flat template's scope — the §5.2 narrowing. The
engine never imports UI code (Iron Rule 1).
"""
from __future__ import annotations

import csv as _csv
import datetime as dt
from dataclasses import dataclass, field
from typing import Optional

from pydantic import ValidationError

from engine.models import (
    Lease,
    LeaseStatus,
    LeaseType,
    MiscItemSpec,
    MoneyRate,
    MoneyUnit,
    RentStep,
    UponExpiration,
)
from engine.models.leases import BASE_RENT_UNITS

RENT_ROLL_SHEET = "Rent Roll"
RENT_STEPS_SHEET = "Rent Steps"
MISC_ITEMS_SHEET = "Misc Items"

#: Columns the Rent Roll sheet must contain (a lease cannot be built
#: without them); the rest are optional per the Lease schema defaults.
_REQUIRED_RENT_ROLL_COLUMNS = [
    "tenant_name", "area", "lease_type", "start_date",
    "base_rent_amount", "base_rent_unit",
]

_BASE_RENT_UNIT_VALUES = [u.value for u in MoneyUnit if u in BASE_RENT_UNITS]


class RentRollImportError(ValueError):
    """Raised when the template has any error. ``errors`` is the full list of
    plain-language, row-level messages; ``str`` joins them for display."""

    def __init__(self, errors: list[str]):
        self.errors = list(errors)
        body = "\n".join(f"  - {e}" for e in self.errors)
        super().__init__(
            "The rent roll template could not be imported "
            f"({len(self.errors)} problem(s) found):\n{body}")


@dataclass
class ImportResult:
    """The outcome of a successful import: the ``leases`` built from the
    Contractual rows, plus readable informational ``notes`` (e.g.
    engine-projected Speculative rows that were intentionally ignored — a
    stated skip, never silent). Iterating an ``ImportResult`` yields its
    leases, and ``len`` is the lease count, so it reads like the list it
    largely is."""

    leases: list[Lease]
    notes: list[str] = field(default_factory=list)

    def __iter__(self):
        return iter(self.leases)

    def __len__(self):
        return len(self.leases)

    def __getitem__(self, index):
        return self.leases[index]


@dataclass
class _Sheet:
    name: str
    columns: list[str]
    rows: list[tuple[int, dict]] = field(default_factory=list)  # (excel_row, values)


# ------------------------------------------------------------------ #
# Readable field coercion (collects errors; never raises per-field)  #
# ------------------------------------------------------------------ #

class _RowErrors:
    """Accumulates readable messages for one sheet, prefixing each with the
    sheet / row / column location."""

    def __init__(self, sheet: str, row_num: int, sink: list[str]):
        self.sheet, self.row_num, self.sink = sheet, row_num, sink
        self.ok = True

    def add(self, column: str, problem: str, fix: str) -> None:
        self.ok = False
        self.sink.append(
            f"{self.sheet} sheet, row {self.row_num}, column '{column}': "
            f"{problem}. {fix}")

    def add_row(self, problem: str, fix: str) -> None:
        self.ok = False
        self.sink.append(
            f"{self.sheet} sheet, row {self.row_num}: {problem}. {fix}")


def _is_blank(value) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def _text(value) -> Optional[str]:
    if _is_blank(value):
        return None
    return str(value).strip()


def _number(err: _RowErrors, column: str, value, *, positive=False,
            integer=False, required=False, example="12000"):
    if _is_blank(value):
        if required:
            err.add(column, "required value is missing",
                    f"Enter a number, e.g. {example}")
        return None
    try:
        num = float(value)
    except (TypeError, ValueError):
        err.add(column, f"{value!r} is not a number",
                f"Enter a number, e.g. {example}")
        return None
    if integer and num != int(num):
        err.add(column, f"{value!r} is not a whole number",
                f"Enter a whole number, e.g. {example}")
        return None
    if positive and num <= 0:
        err.add(column, f"{value!r} is not a positive number",
                f"Enter a value greater than zero, e.g. {example}")
        return None
    return int(num) if integer else num


def _enum(err: _RowErrors, column: str, value, values: list[str], *,
          what: str, required=False):
    if _is_blank(value):
        if required:
            err.add(column, "required value is missing",
                    f"Use one of: {', '.join(values)}")
        return None
    text = str(value).strip()
    if text not in values:
        err.add(column, f"{value!r} is not a valid {what}",
                f"Use one of: {', '.join(values)}")
        return None
    return text


def _date(err: _RowErrors, column: str, value, *, required=False):
    if _is_blank(value):
        if required:
            err.add(column, "required value is missing",
                    "Use the format YYYY-MM-DD, e.g. 2026-01-01")
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    try:
        return dt.date.fromisoformat(str(value).strip())
    except ValueError:
        err.add(column, f"{value!r} is not a valid date",
                "Use the format YYYY-MM-DD, e.g. 2026-01-01")
        return None


def _bool(value) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in ("true", "1", "yes", "y")


# ------------------------------------------------------------------ #
# Reading (xlsx / csv) → sheets of (row_num, {column: value})        #
# ------------------------------------------------------------------ #

def _sheet_from_rows(name: str, raw: list[list]) -> _Sheet:
    if not raw:
        return _Sheet(name=name, columns=[])
    header = [str(c).strip() if c is not None else "" for c in raw[0]]
    sheet = _Sheet(name=name, columns=header)
    for i, row in enumerate(raw[1:]):
        values = {header[j]: (row[j] if j < len(row) else None)
                  for j in range(len(header))}
        if all(_is_blank(v) for v in values.values()):
            continue  # skip fully blank rows
        sheet.rows.append((i + 2, values))  # +2: header is row 1
    return sheet


def _read_xlsx(path) -> dict[str, _Sheet]:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheets = {}
    for name in (RENT_ROLL_SHEET, RENT_STEPS_SHEET, MISC_ITEMS_SHEET):
        if name in wb.sheetnames:
            raw = [list(r) for r in wb[name].iter_rows(values_only=True)]
            sheets[name] = _sheet_from_rows(name, raw)
        else:
            sheets[name] = _Sheet(name=name, columns=[])
    wb.close()
    return sheets


def _read_csv_file(path, name: str) -> _Sheet:
    if path is None:
        return _Sheet(name=name, columns=[])
    with open(path, newline="", encoding="utf-8-sig") as handle:
        raw = [list(r) for r in _csv.reader(handle)]
    return _sheet_from_rows(name, raw)


# ------------------------------------------------------------------ #
# Reconstruction                                                     #
# ------------------------------------------------------------------ #

def _translate_pydantic(exc: ValidationError, sheet: str, row_num: int,
                        errors: list[str]) -> None:
    """Turn a pydantic ``ValidationError`` into plain row-level messages
    (the fallback for cross-field rules — exactly-one-of term/end,
    end<=start, option/reabsorb — that the per-field checks don't cover).
    Our model validators already carry readable text; built-in constraint
    messages are passed through with the field named."""
    for e in exc.errors():
        loc = ".".join(str(p) for p in e.get("loc", ())) or "(row)"
        errors.append(
            f"{sheet} sheet, row {row_num}: {e['msg']} "
            f"(field '{loc}'). Correct this row and re-import.")


def _steps_by_tenant(sheet: _Sheet, errors: list[str]
                     ) -> dict[str, list[RentStep]]:
    out: dict[str, list[RentStep]] = {}
    for row_num, values in sheet.rows:
        err = _RowErrors(sheet.name, row_num, errors)
        tenant = _text(values.get("tenant_name"))
        if tenant is None:
            err.add("tenant_name", "no tenant name",
                    "Every rent-step row must name the tenant it belongs to")
            continue
        amount = _number(err, "amount", values.get("amount"), example="11.00")
        pct = _number(err, "pct_increase", values.get("pct_increase"),
                      example="3.0")
        unit = _enum(err, "unit", values.get("unit"),
                     [u.value for u in MoneyUnit], what="money unit")
        date = _date(err, "date", values.get("date"))
        month_offset = _number(err, "month_offset", values.get("month_offset"),
                               integer=True, example="24")
        if not err.ok:
            continue
        try:
            step = RentStep(
                amount=amount, pct_increase=pct,
                unit=(MoneyUnit(unit) if unit else None),
                date=date, month_offset=month_offset)
        except ValidationError as exc:
            _translate_pydantic(exc, sheet.name, row_num, errors)
            continue
        except ValueError as exc:
            err.add_row(f"this rent-step row could not be read ({exc})",
                        "Check the values in this row and re-import")
            continue
        out.setdefault(tenant, []).append(step)
    return out


def _misc_by_tenant(sheet: _Sheet, errors: list[str]
                    ) -> dict[str, list[MiscItemSpec]]:
    out: dict[str, list[MiscItemSpec]] = {}
    for row_num, values in sheet.rows:
        err = _RowErrors(sheet.name, row_num, errors)
        tenant = _text(values.get("tenant_name"))
        if tenant is None:
            err.add("tenant_name", "no tenant name",
                    "Every misc-item row must name the tenant it belongs to")
            continue
        name = _text(values.get("name"))
        if name is None:
            err.add("name", "required value is missing",
                    "Enter a name, e.g. 'Storage'")
        amount = _number(err, "amount", values.get("amount"), required=True,
                         example="100")
        unit = _enum(err, "unit", values.get("unit"),
                     [u.value for u in MoneyUnit], what="money unit",
                     required=True)
        if not err.ok:
            continue
        try:
            item = MiscItemSpec(name=name, amount=amount, unit=MoneyUnit(unit),
                                free_rent_abates=_bool(
                                    values.get("free_rent_abates")))
        except ValidationError as exc:
            _translate_pydantic(exc, sheet.name, row_num, errors)
            continue
        except ValueError as exc:
            err.add_row(f"this misc-item row could not be read ({exc})",
                        "Check the values in this row and re-import")
            continue
        out.setdefault(tenant, []).append(item)
    return out


def _build_leases(sheets: dict[str, _Sheet]) -> list[Lease]:
    errors: list[str] = []
    roll = sheets[RENT_ROLL_SHEET]

    if not roll.columns:
        raise RentRollImportError([
            f"The '{RENT_ROLL_SHEET}' sheet is missing or empty. It must have "
            "a header row and one row per lease."])
    missing = [c for c in _REQUIRED_RENT_ROLL_COLUMNS if c not in roll.columns]
    if missing:
        raise RentRollImportError([
            f"{RENT_ROLL_SHEET} sheet: required column '{c}' is missing. "
            f"Add a '{c}' column to the header row." for c in missing])

    steps = _steps_by_tenant(sheets[RENT_STEPS_SHEET], errors)
    misc = _misc_by_tenant(sheets[MISC_ITEMS_SHEET], errors)

    leases: list[Lease] = []
    seen_tenants: dict[str, int] = {}
    speculative_rows: list[int] = []
    for row_num, values in roll.rows:
        # Provenance filter: only Contractual rows become leases. Speculative
        # rows are engine projections (MLP rollover / absorption), NOT intake
        # — ignored for lease construction but recorded and reported (not a
        # silent skip; owner directive). A blank provenance stays Contractual
        # (a plain rent roll with no ``status`` column imports fully, as in
        # Step 7); anything not exactly "Speculative" is treated as
        # Contractual (the label is a soft presentation hint, not an enum).
        provenance = _text(values.get("status"))
        if provenance is not None and provenance.strip().lower() == "speculative":
            speculative_rows.append(row_num)
            continue

        err = _RowErrors(RENT_ROLL_SHEET, row_num, errors)
        tenant = _text(values.get("tenant_name"))
        if tenant is None:
            err.add("tenant_name", "no tenant name",
                    "Enter the tenant's name, e.g. 'Acme Co'")
        elif tenant in seen_tenants:
            err.add("tenant_name",
                    f"tenant name {tenant!r} is already used on row "
                    f"{seen_tenants[tenant]}",
                    "Tenant names must be unique — rename one of the leases")
        else:
            seen_tenants[tenant] = row_num

        # Required fields (missing/blank AND wrong-typed both route through
        # the readable RentRollImportError — never a raw pydantic/ValueError
        # dump; owner review of Step 7, DEVIATIONS.md §25 discrimination).
        area = _number(err, "area", values.get("area"), positive=True,
                       required=True, example="12000")
        lease_type = _enum(err, "lease_type", values.get("lease_type"),
                           [t.value for t in LeaseType], what="lease type",
                           required=True)
        base_amount = _number(err, "base_rent_amount",
                              values.get("base_rent_amount"), required=True,
                              example="25.00")
        base_unit = _enum(err, "base_rent_unit", values.get("base_rent_unit"),
                          _BASE_RENT_UNIT_VALUES, what="base rent unit",
                          required=True)
        start_date = _date(err, "start_date", values.get("start_date"),
                           required=True)
        # Optional: a blank cell means "use the §3 schema default"
        # (``lease_status`` → contract, ``upon_expiration`` → market) or the
        # one-of pair (``end_date`` / ``term_months``, checked by the Lease
        # validator). None of these leak — see the round-trip default tests.
        # ``lease_status`` is the §3.12 status (contract/mtm/speculative),
        # distinct from the provenance ``status`` column filtered above.
        status = _enum(err, "lease_status", values.get("lease_status"),
                       [s.value for s in LeaseStatus], what="lease status")
        upon = _enum(err, "upon_expiration", values.get("upon_expiration"),
                     [u.value for u in UponExpiration],
                     what="upon-expiration option")
        end_date = _date(err, "end_date", values.get("end_date"))
        term_months = _number(err, "term_months", values.get("term_months"),
                              integer=True, positive=True, example="120")

        if not err.ok:
            continue

        # Build the nested value objects and the Lease inside one translating
        # guard: after the required checks above nothing here should raise on
        # a blank, but the enum/MoneyRate constructors raise a raw ValueError
        # on any unexpected input — catch BOTH so no raw error can escape an
        # intake surface (spec §5.4).
        try:
            kwargs = dict(
                tenant_name=tenant,
                suite=_text(values.get("suite")),
                external_id=_text(values.get("external_id")),
                area=area,
                lease_type=LeaseType(lease_type),
                start_date=start_date,
                base_rent=MoneyRate(amount=base_amount,
                                    unit=MoneyUnit(base_unit)),
                market_leasing_profile=_text(
                    values.get("market_leasing_profile")),
                notes=_text(values.get("notes")),
                rent_steps=steps.get(tenant, []),
                miscellaneous_items=misc.get(tenant, []),
            )
            if end_date is not None:
                kwargs["end_date"] = end_date
            if term_months is not None:
                kwargs["term_months"] = term_months
            if status is not None:
                kwargs["status"] = LeaseStatus(status)
            if upon is not None:
                kwargs["upon_expiration"] = UponExpiration(upon)
            leases.append(Lease(**kwargs))
        except ValidationError as exc:
            _translate_pydantic(exc, RENT_ROLL_SHEET, row_num, errors)
        except ValueError as exc:
            err.add_row(f"this row could not be read ({exc})",
                        "Check the values in this row and re-import")

    if errors:
        raise RentRollImportError(errors)
    notes: list[str] = []
    if speculative_rows:
        rows_text = ", ".join(str(r) for r in speculative_rows)
        notes.append(
            f"{len(speculative_rows)} speculative/projected row(s) ignored — "
            f"engine projections (MLP rollover / absorption), not intake "
            f"(Rent Roll rows {rows_text}).")
    return ImportResult(leases=leases, notes=notes)


# ------------------------------------------------------------------ #
# Public front-ends                                                  #
# ------------------------------------------------------------------ #

def import_rent_roll(path) -> ImportResult:
    """Import an ``.xlsx`` rent-roll template (Rent Roll + Rent Steps + Misc
    Items sheets) into an :class:`ImportResult` — the validated leases (the
    Contractual rows) plus readable ``notes`` (e.g. Speculative rows
    ignored). Raises :class:`RentRollImportError` (with plain, row-level
    messages) if any Contractual row is invalid."""
    return _build_leases(_read_xlsx(path))


def import_rent_roll_csv(rent_roll_path, *, steps_path=None,
                         misc_path=None) -> ImportResult:
    """Import the rent roll from CSV (spec §5.2 "also support CSV"): the Rent
    Roll rows in ``rent_roll_path``, with optional companion CSVs for rent
    steps and misc items. Same validation, readable errors, and
    :class:`ImportResult` as the xlsx path."""
    sheets = {
        RENT_ROLL_SHEET: _read_csv_file(rent_roll_path, RENT_ROLL_SHEET),
        RENT_STEPS_SHEET: _read_csv_file(steps_path, RENT_STEPS_SHEET),
        MISC_ITEMS_SHEET: _read_csv_file(misc_path, MISC_ITEMS_SHEET),
    }
    return _build_leases(sheets)
